""" Collection of functions for exporting data.
    Requires openpyxl for Excel exports
"""
import csv

from openpyxl import Workbook
from openpyxl.cell import get_column_letter


def write_to_file(filename, data_sample):
    """ Write the entire data sample to a pdf """
    example = csv.writer(open(filename, 'w', newline='', encoding='utf-8'))
    example.writerows(data_sample)

# csv with sorted column
# ascending is going up i.e. A-Z 
# descending is going down i.e. Z-A 
def write_sorted_col_to_pdf(filename, data_sample, col, order="ascending"):
    """ Orders and write the 1 column of data sample to a pdf """
    if order == "descending":
        data_sample.sort(key=lambda x: float(x[col]), reverse=False)
    else:
        data_sample.sort(key=lambda x: float(x[col]), reverse=True)
    write_to_file(filename, data_sample) 


# export to excel
def save_spreadsheet(filename, data_sample): 
    """ Save data sample to an Excel sheet """  
    wb = Workbook()
    ws = wb.active
    rowIndex = 1
    for rows in data_sample:
        colIndex = 1 
        for field in rows:
            colIndex2 = get_column_letter(colIndex)
            ws.cell('%s%s'%(colIndex2, rowIndex)).value = field
            colIndex +=1
        rowIndex += 1 
    wb.save(filename)